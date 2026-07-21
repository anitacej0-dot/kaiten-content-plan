"""Загрузка входных сигналов цикла.

Источники:
  * Roistat-выгрузка (лист «Report», «Посадочная страница» + Визиты/Заявки/Продажи)
      -> лид-потенциал и Protect по страницам блога;
  * Метрика «Посещаемость» (лист «Отчет», дневные визиты) -> общий тренд трафика;
  * текстовые/таблич. шаблоны редакции в inputs/current_cycle/:
      seo_signals.csv, editorial_ideas.csv, mandatory_publications.csv,
      product_priorities.md, sales_kam_signals.md, cases_and_experts.md,
      editorial_capacity.md, launches_and_deadlines.md.

Файлы распознаются по СОДЕРЖИМОМУ (сигнатуре заголовков), а не только по имени,
поэтому реальные выгрузки Roistat/Метрики подхватываются даже под своими именами.
Отсутствие любого необязательного источника не роняет pipeline — фиксируется пропуск.
"""
from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional

from openpyxl import load_workbook

from .io_utils import find_files, read_csv, read_text
from .warnings_log import WarningLog

# --------------------------------------------------------------- транслитерация
_TRANSLIT = {
    "а": "a", "б": "b", "в": "v", "г": "g", "д": "d", "е": "e", "ё": "e",
    "ж": "zh", "з": "z", "и": "i", "й": "y", "к": "k", "л": "l", "м": "m",
    "н": "n", "о": "o", "п": "p", "р": "r", "с": "s", "т": "t", "у": "u",
    "ф": "f", "х": "h", "ц": "c", "ч": "ch", "ш": "sh", "щ": "sch", "ъ": "",
    "ы": "y", "ь": "", "э": "e", "ю": "yu", "я": "ya", " ": "-",
}


def translit(text: str) -> str:
    text = (text or "").lower().replace("ё", "е")
    out = []
    for ch in text:
        if ch in _TRANSLIT:
            out.append(_TRANSLIT[ch])
        elif ch.isalnum():
            out.append(ch)
        else:
            out.append("-")
    s = "".join(out)
    return re.sub(r"-+", "-", s).strip("-")


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(str(v).replace(",", ".").replace(" ", ""))
    except (ValueError, TypeError):
        return 0.0


# --------------------------------------------------------------- Roistat
@dataclass
class PageSignal:
    url: str
    slug: str
    visits: float = 0.0
    leads: float = 0.0
    sales: float = 0.0
    revenue: float = 0.0


def _detect_kind(ws) -> str:
    """Определить тип выгрузки по первым строкам: roistat | metrika | other."""
    try:
        first = [str(c.value).strip().lower() if c.value is not None else ""
                 for c in next(ws.iter_rows(min_row=1, max_row=1))]
    except StopIteration:
        return "other"
    joined = " ".join(first)
    if "посадочная страница" in joined:
        return "roistat"
    if "отчет за период" in joined or "визиты" in joined:
        return "metrika"
    return "other"


def load_roistat(paths: List[str], warn: WarningLog) -> Dict:
    """Агрегировать Roistat по посадочным страницам (по всем найденным файлам)."""
    pages: Dict[str, PageSignal] = {}
    used_files: List[str] = []
    for path in paths:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            warn.warn("signals.roistat", f"Не открыть {os.path.basename(path)}: {e}")
            continue
        ws = wb[wb.sheetnames[0]]
        if _detect_kind(ws) != "roistat":
            wb.close()
            continue
        used_files.append(os.path.basename(path))
        rows = ws.iter_rows(values_only=True)
        header = next(rows, None)
        idx = _roistat_columns(header)
        for r in rows:
            if not r or idx["url"] >= len(r):
                continue
            url = str(r[idx["url"]] or "").strip()
            if not url or "kaiten" not in url.lower():
                continue
            key = url.lower().split("?")[0].rstrip("/")
            ps = pages.get(key)
            if ps is None:
                ps = PageSignal(url=key, slug=key.split("/")[-1])
                pages[key] = ps
            ps.visits += _num(r[idx["visits"]]) if idx["visits"] < len(r) else 0
            ps.leads += _num(r[idx["leads"]]) if idx["leads"] < len(r) else 0
            ps.sales += _num(r[idx["sales"]]) if idx["sales"] < len(r) else 0
            ps.revenue += _num(r[idx["revenue"]]) if idx["revenue"] < len(r) else 0
        wb.close()
    plist = sorted(pages.values(), key=lambda p: (p.leads, p.visits), reverse=True)
    totals = {
        "pages": len(plist),
        "visits": sum(p.visits for p in plist),
        "leads": sum(p.leads for p in plist),
        "sales": sum(p.sales for p in plist),
        "files": used_files,
    }
    if used_files:
        warn.info("signals.roistat",
                  f"Roistat: {len(plist)} страниц, заявок {int(totals['leads'])}, "
                  f"визитов {int(totals['visits'])} ({', '.join(used_files)})")
    return {"pages": plist, "totals": totals}


def _roistat_columns(header) -> Dict[str, int]:
    """Найти индексы нужных колонок Roistat по названиям (устойчиво к сдвигам)."""
    names = [str(h).strip().lower() if h is not None else "" for h in (header or [])]

    def find(*keys, default):
        for i, n in enumerate(names):
            if any(k in n for k in keys):
                return i
        return default
    return {
        "url": find("посадочная страница", default=0),
        "visits": find("визиты", default=15),
        "leads": find("заявк", default=16),
        "sales": find("продажи", default=18),
        "revenue": find("выручка", default=19),
    }


# --------------------------------------------------------------- Метрика
def load_metrika(paths: List[str], warn: WarningLog) -> Dict:
    for path in paths:
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:
            warn.warn("signals.metrika", f"Не открыть {os.path.basename(path)}: {e}")
            continue
        ws = wb[wb.sheetnames[0]]
        if _detect_kind(ws) != "metrika":
            wb.close()
            continue
        daily = []
        for r in ws.iter_rows(values_only=True):
            if not r or len(r) < 2:
                continue
            d, v = r[0], r[1]
            if isinstance(d, str) and re.match(r"\d{4}-\d{2}-\d{2}", d.strip()):
                daily.append(_num(v))
            elif hasattr(d, "strftime"):
                daily.append(_num(v))
        wb.close()
        if daily:
            total = sum(daily)
            avg = total / len(daily)
            n = len(daily)
            half = max(1, n // 2)
            trend = 0.0
            if sum(daily[:half]) > 0:
                trend = (sum(daily[half:]) - sum(daily[:half])) / sum(daily[:half]) * 100
            warn.info("signals.metrika",
                      f"Метрика: {n} дней, всего визитов {int(total)}, "
                      f"средн/день {int(avg)}, тренд {trend:+.0f}%")
            return {"days": n, "total": total, "avg": avg, "trend_pct": trend,
                    "file": os.path.basename(path)}
    return {}


# --------------------------------------------------------------- текст. сигналы
# Ожидаемые файлы редакции (имя -> человекочитаемая подпись).
EXPECTED_TEXT_SIGNALS = {
    "seo_signals.csv": "SEO-сигналы",
    "editorial_ideas.csv": "Идеи редакции",
    "mandatory_publications.csv": "Обязательные публикации",
    "product_priorities.md": "Продуктовые приоритеты",
    "sales_kam_signals.md": "Запросы продаж / KAM",
    "cases_and_experts.md": "Кейсы и эксперты",
    "editorial_capacity.md": "Мощность редакции",
    "launches_and_deadlines.md": "Запуски и дедлайны",
}


@dataclass
class Signals:
    roistat: Dict = field(default_factory=dict)
    metrika: Dict = field(default_factory=dict)
    csv_rows: Dict[str, List[Dict]] = field(default_factory=dict)   # имя файла -> строки
    text_files: Dict[str, str] = field(default_factory=dict)        # имя файла -> текст
    present: List[str] = field(default_factory=list)
    missing: List[str] = field(default_factory=list)

    def has(self, name: str) -> bool:
        return name in self.csv_rows or name in self.text_files


def load_signals(cfg, warn: WarningLog) -> Signals:
    directory = cfg.inputs_current
    sig = Signals()
    # Все xlsx в каталоге и подкаталоге analytics/
    xlsx_paths = find_files(directory, [".xlsx"])
    xlsx_paths += find_files(os.path.join(directory, "analytics"), [".xlsx"])
    sig.roistat = load_roistat(xlsx_paths, warn)
    sig.metrika = load_metrika(xlsx_paths, warn)
    if sig.roistat.get("totals", {}).get("pages"):
        sig.present.append("Roistat (заявки/трафик по страницам)")
    else:
        sig.missing.append("Roistat (заявки/трафик по страницам)")
    if sig.metrika:
        sig.present.append("Метрика (тренд посещаемости)")
    else:
        sig.missing.append("Метрика (тренд посещаемости)")

    for fname, label in EXPECTED_TEXT_SIGNALS.items():
        path = os.path.join(directory, fname)
        if not os.path.exists(path):
            sig.missing.append(label)
            continue
        if fname.endswith(".csv"):
            rows = [r for r in read_csv(path) if any((v or "").strip() for v in r.values())]
            sig.csv_rows[fname] = rows
            sig.present.append(f"{label} ({len(rows)} строк)")
        else:
            txt = read_text(path).strip()
            if txt:
                sig.text_files[fname] = txt
                sig.present.append(label)
            else:
                sig.missing.append(label)
    return sig
