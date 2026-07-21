"""Разбор Excel-книги контент-плана в единую историческую схему.

Особенности реального файла (учтены):
  * 39 листов: месяцы блога, внешние площадки, кейсы, пустой Лист11;
  * месяц/год и канал закодированы в имени листа: «Февраль 25», «Май 2025 (блог)»,
    «Февраль 25 (VC)», «Апрель 25 (внешние)», «Кейсы»;
  * заголовки с хвостовыми пробелами и вариациями («ТЗ SEO» / «ТЗ от Максима»,
    «Кто пишет» / «С кем пишем»);
  * внешние листы «раздуты» форматированием до ~1000 строк — реальных строк единицы;
  * площадка иногда в колонке «Площадка», иногда в имени листа ((VC)/(Habr)).

Ошибка на одном листе не роняет разбор остальных.
"""
from __future__ import annotations

import datetime as _dt
import os
import re
from typing import Dict, List, Optional

from openpyxl import load_workbook

from .taxonomy import Classifier
from .text_utils import TextNormalizer, contains_any
from .warnings_log import WarningLog

# Нормализованный заголовок -> канонический ключ.
HEADER_MAP = {
    "тема": "original_title",
    "описание": "description",
    "рубрика": "rubric",
    "раздел": "section",
    "дата": "date_planned",
    "публикация по факту": "date_fact",
    "необходимость эксперта": "expert_need",
    "тз seo": "seo_tz",
    "готовое тз": "ready_tz",
    "комментарий": "comment",
    "комментарии от даши": "comment2",
    "статус": "status",
    "тз от максима": "seo_keyword",
    "кто пишет": "author",
    "с кем пишем": "coauthor",
    "площадка": "platform_raw",
    "карточка в kaiten": "kaiten_card",
    "куда пойдет": "destination",
    "когда планируем выпуск": "when_planned",
}

MONTHS = {
    "январ": 1, "феврал": 2, "март": 3, "апрел": 4, "мая": 5, "май": 5,
    "июн": 6, "июл": 7, "август": 8, "сентябр": 9, "октябр": 10,
    "ноябр": 11, "декабр": 12,
}

_REWRITE_MARKERS = ["рерайт", "обновлен", "обновить", "что нового", "дайджест",
                    "апдейт", "актуализац", "переработ"]


def _norm_header(h) -> str:
    if h is None:
        return ""
    return re.sub(r"\s+", " ", str(h).strip().lower().replace("ё", "е"))


def _map_header(h: str) -> Optional[str]:
    nh = _norm_header(h)
    if not nh:
        return None
    if nh in HEADER_MAP:
        return HEADER_MAP[nh]
    for key, canon in HEADER_MAP.items():
        if nh.startswith(key) or key in nh:
            return canon
    return None


def parse_period(sheet_name: str) -> tuple:
    """('2025-02', 'Февраль 2025') из имени листа; ('', '') если не найдено."""
    low = sheet_name.lower().replace("ё", "е")
    month = None
    for stem, num in MONTHS.items():
        if stem in low:
            month = num
            break
    ym = re.search(r"(20\d{2}|\b\d{2}\b)", sheet_name)
    year = None
    if ym:
        y = ym.group(1)
        year = int(y) if len(y) == 4 else 2000 + int(y)
    if month and year:
        return f"{year:04d}-{month:02d}", f"{_month_name(month)} {year}"
    return "", ""


def _month_name(num: int) -> str:
    names = ["", "Январь", "Февраль", "Март", "Апрель", "Май", "Июнь", "Июль",
             "Август", "Сентябрь", "Октябрь", "Ноябрь", "Декабрь"]
    return names[num] if 1 <= num <= 12 else ""


def classify_sheet(sheet_name: str, cfg_legacy: Dict) -> str:
    low = sheet_name.lower().replace("ё", "е")
    for marker in cfg_legacy.get("ignore_sheet_markers", []):
        if low.strip() == marker or (marker == "лист" and low.startswith("лист")):
            return "ignore"
    if any(m in low for m in cfg_legacy.get("case_sheet_markers", ["кейс"])):
        return "case"
    if "(vc)" in low or "(habr)" in low or "внешн" in low:
        return "external"
    return "blog"


def _to_date_str(value) -> str:
    if value is None or value == "":
        return ""
    if isinstance(value, (_dt.datetime, _dt.date)):
        return value.strftime("%Y-%m-%d")
    s = str(value).strip()
    return s.split(" ")[0] if s else ""


class LegacyParser:
    def __init__(self, config, warnings: Optional[WarningLog] = None):
        self.cfg = config
        self.legacy_cfg = config.pipeline.get("legacy_plan", {})
        self.classifier = Classifier(config.taxonomy)
        self.normalizer = TextNormalizer(config.taxonomy)
        self.warn = warnings or WarningLog()
        self.min_title_len = int(self.legacy_cfg.get("min_title_len", 3))

    def parse(self, path: str) -> List[Dict]:
        if not path or not os.path.exists(path):
            self.warn.error("parse_legacy", f"Файл не найден: {path}")
            return []
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
        except Exception as e:  # повреждённый/занятый файл
            self.warn.error("parse_legacy", f"Не удалось открыть {os.path.basename(path)}: {e}")
            return []

        rows: List[Dict] = []
        rid = 0
        for sheet_name in wb.sheetnames:
            kind = classify_sheet(sheet_name, self.legacy_cfg)
            if kind == "ignore":
                self.warn.info("parse_legacy", f"Лист пропущен: '{sheet_name}'")
                continue
            try:
                sheet_rows = self._parse_sheet(wb[sheet_name], sheet_name, kind, rid)
                rid += len(sheet_rows)
                rows.extend(sheet_rows)
            except Exception as e:
                # один плохой лист не должен ронять весь разбор
                self.warn.warn("parse_legacy",
                               f"Лист '{sheet_name}' пропущен из-за ошибки: {e}")
        wb.close()
        self.warn.info("parse_legacy",
                       f"Разобрано строк истории: {len(rows)} из {len(wb.sheetnames)} листов")
        return rows

    def _parse_sheet(self, ws, sheet_name: str, kind: str, start_rid: int) -> List[Dict]:
        period, period_label = parse_period(sheet_name)
        # Прочитать все строки один раз (read_only отдаёт генератор).
        all_rows = list(ws.iter_rows(values_only=True))
        if not all_rows:
            return []
        header = all_rows[0]
        col_map: Dict[int, str] = {}
        for idx, cell in enumerate(header):
            canon = _map_header(cell)
            if canon:
                col_map.setdefault(idx, canon)
        # На листах кейсов темы часто нет: название кейса лежит в «Описание».
        if "original_title" not in col_map.values() and kind == "case":
            for idx, canon in list(col_map.items()):
                if canon == "description":
                    col_map[idx] = "original_title"
                    break
        if "original_title" not in col_map.values():
            self.warn.warn("parse_legacy",
                           f"Лист '{sheet_name}': не найдена колонка «Тема», пропущен")
            return []

        out: List[Dict] = []
        empty_streak = 0
        rid = start_rid
        for r in all_rows[1:]:
            record = {canon: r[idx] for idx, canon in col_map.items() if idx < len(r)}
            title = str(record.get("original_title") or "").strip()
            if len(title) < self.min_title_len:
                empty_streak += 1
                if empty_streak >= 30:  # хвост из пустых строк -> заканчиваем лист
                    break
                continue
            empty_streak = 0
            rid += 1
            out.append(self._build_row(rid, sheet_name, kind, period, period_label, record))
        return out

    def _build_row(self, rid, sheet_name, kind, period, period_label, rec) -> Dict:
        title = str(rec.get("original_title") or "").strip()
        desc = str(rec.get("description") or "").strip()
        rubric = str(rec.get("rubric") or "").strip()
        section = str(rec.get("section") or "").strip()
        status = str(rec.get("status") or "").strip()

        # площадка / канал
        if kind == "case":
            channel = "case"
            platform_label, _pk = self.classifier.platform(
                str(rec.get("platform_raw") or rec.get("destination") or ""), sheet_name)
        elif kind == "external":
            channel = "external"
            platform_label, _pk = self.classifier.platform(
                str(rec.get("platform_raw") or ""), sheet_name)
        else:
            channel = "blog"
            platform_label = "Блог Kaiten"

        ctype_key, ctype_label, _ = self.classifier.content_type(title, desc, rubric, section)
        cluster_label, _ = self.classifier.product_cluster(title, desc, rubric, section)
        role_label, _ = self.classifier.strategy_role(title, desc, rubric, section, ctype_key)
        # Кейсы всегда доказательный контент.
        if kind == "case":
            ctype_key, ctype_label = "case", "Кейс"
            role_label = "Prove"

        comment = " | ".join(x for x in [str(rec.get("comment") or "").strip(),
                                         str(rec.get("comment2") or "").strip()] if x)
        seo_tz = str(rec.get("seo_tz") or "").strip()
        seo_keyword = str(rec.get("seo_keyword") or "").strip()
        # «ТЗ от Максима» часто содержит SEO-ключ/заголовок — это SEO-сигнал.
        if seo_keyword and seo_keyword.lower() not in ("без тз на seo", "без тз по seo", "нет"):
            seo_tz = (seo_tz + " | " + seo_keyword).strip(" |")

        is_done = self.classifier.is_done(status)
        is_rewrite = contains_any(title + " " + desc, _REWRITE_MARKERS)

        return {
            "row_id": f"H{rid:04d}",
            "source_sheet": sheet_name,
            "sheet_kind": kind,
            "period": period,
            "period_label": period_label,
            "channel": channel,
            "platform": platform_label,
            "original_title": title,
            "norm_title": self.normalizer.norm_title(title),
            "description": desc,
            "rubric": rubric,
            "section": section,
            "date_planned": _to_date_str(rec.get("date_planned")),
            "date_fact": _to_date_str(rec.get("date_fact")),
            "author": str(rec.get("author") or "").strip(),
            "coauthor": str(rec.get("coauthor") or "").strip(),
            "seo_tz": seo_tz,
            "ready_tz": str(rec.get("ready_tz") or "").strip(),
            "comment": comment,
            "status": status,
            "kaiten_card": str(rec.get("kaiten_card") or "").strip(),
            "product_cluster": cluster_label,
            "strategy_role": role_label,
            "content_type": ctype_label,
            "is_done": "да" if is_done else "нет",
            "is_rewrite": "да" if is_rewrite else "нет",
        }


def find_legacy_file(directory: str, globs: List[str], warnings: WarningLog) -> Optional[str]:
    """Найти файл легаси-плана в каталоге по маскам. Берём самый свежий .xlsx."""
    import fnmatch
    if not os.path.isdir(directory):
        return None
    candidates: List[str] = []
    names = [n for n in os.listdir(directory)
             if n.lower().endswith(".xlsx") and not n.startswith("~$")]
    for pattern in globs:
        for n in names:
            if fnmatch.fnmatch(n.lower(), pattern.lower()) and n not in candidates:
                candidates.append(n)
    if not candidates:
        return None
    # приоритет — с «контент» в имени, затем по времени изменения
    def sort_key(n):
        has_kw = 1 if ("контент" in n.lower() or "content" in n.lower()) else 0
        mtime = os.path.getmtime(os.path.join(directory, n))
        return (has_kw, mtime)
    best = sorted(candidates, key=sort_key, reverse=True)[0]
    return os.path.join(directory, best)
