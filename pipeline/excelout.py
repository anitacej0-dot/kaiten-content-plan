"""Экспорт итогового контент-плана в .xlsx (11 листов, рабочее форматирование).

Листы: Контент-план, Размещения, Бэклог, Обновления, Кейсы, Отклонённые темы,
Дубли, Нагрузка, Источники сигналов, Decision log, Справочники.

Форматирование: закреплённая шапка, автофильтр, читаемые ширины, перенос строк,
выпадающие списки (валидация), условное форматирование score/обязательных/рисков.
Интерфейс на русском. Справочники используются как источник выпадающих списков —
значит формулы валидации работают сразу после открытия файла.
"""
from __future__ import annotations

from typing import Dict, List, Tuple

from openpyxl import Workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from . import schema

# ---------------------------------------------------------------- стили
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
CELL_FONT = Font(size=10)
THIN = Side(style="thin", color="D9D9D9")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
WRAP_TOP = Alignment(wrap_text=True, vertical="top")
HEADER_ALIGN = Alignment(wrap_text=True, vertical="center", horizontal="center")

FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FILL_MANDATORY = PatternFill("solid", fgColor="FCE4D6")
FILL_RISK = PatternFill("solid", fgColor="FFD0D0")

# Колонки листа «Контент-план»: (key, label, width, wrap)
PLAN_COLUMNS: List[Tuple[str, str, int, bool]] = [
    ("content_id", "content_id", 14, False),
    ("title", "Тема", 40, True),
    ("description", "Краткое описание", 44, True),
    ("strategy_role", "Роль", 10, False),
    ("content_type", "Тип материала", 18, True),
    ("product_cluster", "Кластер", 22, True),
    ("product_feature", "Продукт/функция", 18, True),
    ("segment", "Сегмент", 20, True),
    ("audience_role", "Роль читателя", 18, True),
    ("funnel_stage", "Воронка", 18, True),
    ("user_problem", "Проблема пользователя", 34, True),
    ("idea_source", "Источник идеи", 18, True),
    ("signal", "Сигнал", 30, True),
    ("expected_action", "Ожидаемое действие", 26, True),
    ("cta", "CTA", 22, True),
    ("product_route", "Маршрут к продукту", 22, True),
    ("publication_kind", "Вид", 14, False),
    ("related_content", "Связь с материалами", 30, True),
    ("lead_potential", "Лид-потенциал", 13, False),
    ("seo_potential", "SEO", 10, False),
    ("product_priority", "Приоритет", 12, False),
    ("evidence_base", "Доказательная база", 24, True),
    ("effort", "Трудоёмкость", 12, False),
    ("dependencies", "Зависимости", 18, True),
    ("expert", "Эксперт", 16, True),
    ("author", "Автор", 14, True),
    ("target_week", "Неделя выпуска", 20, True),
    ("score", "Score", 8, False),
    ("rationale", "Обоснование", 40, True),
    ("mandatory", "Обязат.", 9, False),
    ("risk_flags", "Риски", 20, True),
    ("decision_status", "Статус решения", 15, False),
    ("kaiten_card", "Карточка Kaiten", 24, True),
    ("editor_comment", "Коммент. главреда", 30, True),
]

PLACEMENT_COLUMNS = [(k, lbl, w, wr) for (k, lbl), (w, wr) in zip(
    schema.PLACEMENT_FIELDS,
    [(14, False), (14, False), (16, True), (16, True), (34, True), (22, True),
     (24, True), (22, True), (18, False), (16, False), (30, True), (30, True),
     (14, True), (24, True)])]

DUP_COLUMNS = [
    ("group_id", "Группа", 10, False), ("relationship", "Связь", 16, True),
    ("similarity", "Схожесть %", 11, False),
    ("theme_a", "Тема A", 34, True), ("platform_a", "Площадка A", 14, True),
    ("period_a", "Период A", 11, False), ("sheet_a", "Лист A", 18, True),
    ("theme_b", "Тема B", 34, True), ("platform_b", "Площадка B", 14, True),
    ("period_b", "Период B", 11, False), ("sheet_b", "Лист B", 18, True),
    ("note", "Пояснение", 40, True),
]


def _write_table(ws, rows: List[Dict], columns, freeze="A2", wrap_all=False):
    labels = [c[1] for c in columns]
    ws.append(labels)
    for i, (key, label, width, wrap) in enumerate(columns, start=1):
        col = get_column_letter(i)
        ws.column_dimensions[col].width = width
        cell = ws.cell(row=1, column=i)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border = BORDER
    for r in rows:
        ws.append([_fmt(r.get(key, "")) for key, _, _, _ in columns])
    # стиль ячеек данных
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=len(columns)):
        for i, cell in enumerate(row):
            cell.font = CELL_FONT
            cell.border = BORDER
            cell.alignment = WRAP_TOP if (columns[i][3] or wrap_all) else Alignment(vertical="top")
    ws.freeze_panes = freeze
    if ws.max_row >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{max(ws.max_row,1)}"
    ws.sheet_view.showGridLines = True


def _fmt(v):
    if v is None:
        return ""
    if isinstance(v, float):
        return int(v) if float(v).is_integer() else round(v, 1)
    if isinstance(v, (list, tuple, set)):
        return "; ".join(str(x) for x in v if x)
    return v


def _col_index(columns, key):
    for i, c in enumerate(columns, start=1):
        if c[0] == key:
            return i
    return None


def build_workbook(data: Dict) -> Workbook:
    """data: {selected, placements, backlog, updates, cases, rejected, duplicates,
             capacity_lines, signal_rows, decision_log, cycle}"""
    wb = Workbook()

    # ---- Справочники (сначала: на них ссылаются выпадающие списки) ----
    ref = wb.active
    ref.title = "Справочники"
    _build_reference_sheet(ref, data.get("platforms"))

    # ---- Контент-план ----
    ws_plan = wb.create_sheet("Контент-план")
    _write_table(ws_plan, data.get("selected", []), PLAN_COLUMNS, freeze="C2")
    _apply_plan_formatting(ws_plan, len(data.get("selected", [])))

    # ---- Размещения ----
    ws_pl = wb.create_sheet("Размещения")
    _write_table(ws_pl, data.get("placements", []), PLACEMENT_COLUMNS, freeze="C2")
    _apply_placement_validation(ws_pl, len(data.get("placements", [])))

    # ---- Бэклог ----
    ws_bl = wb.create_sheet("Бэклог")
    _write_table(ws_bl, data.get("backlog", []), PLAN_COLUMNS, freeze="C2")
    _apply_plan_formatting(ws_bl, len(data.get("backlog", [])))

    # ---- Обновления (Protect) ----
    ws_up = wb.create_sheet("Обновления")
    _write_table(ws_up, data.get("updates", []), PLAN_COLUMNS, freeze="C2")
    _apply_plan_formatting(ws_up, len(data.get("updates", [])))

    # ---- Кейсы ----
    ws_case = wb.create_sheet("Кейсы")
    _write_table(ws_case, data.get("cases", []), PLAN_COLUMNS, freeze="C2")

    # ---- Отклонённые темы ----
    ws_rej = wb.create_sheet("Отклонённые темы")
    _write_table(ws_rej, data.get("rejected", []), PLAN_COLUMNS, freeze="C2")

    # ---- Дубли ----
    ws_dup = wb.create_sheet("Дубли")
    _write_table(ws_dup, data.get("duplicates", []), DUP_COLUMNS, freeze="A2")
    _apply_dup_formatting(ws_dup, len(data.get("duplicates", [])))

    # ---- Нагрузка ----
    ws_load = wb.create_sheet("Нагрузка")
    _build_load_sheet(ws_load, data.get("capacity_lines", []), data.get("cycle_info", {}))

    # ---- Источники сигналов ----
    ws_sig = wb.create_sheet("Источники сигналов")
    sig_cols = [("source", "Источник", 40, True), ("status", "Статус", 16, True),
                ("detail", "Детали", 60, True)]
    _write_table(ws_sig, data.get("signal_rows", []), sig_cols, freeze="A2")

    # ---- Decision log ----
    ws_dl = wb.create_sheet("Decision log")
    dl_cols = [("date", "Дата", 14, False), ("author", "Кто решил", 18, True),
               ("content_id", "content_id", 14, False), ("field", "Что изменено", 20, True),
               ("old_value", "Было", 26, True), ("new_value", "Стало", 26, True),
               ("reason", "Причина", 40, True), ("is_exception", "Исключение из score", 16, False)]
    _write_table(ws_dl, data.get("decision_log", []), dl_cols, freeze="A2")

    _reorder(wb)
    return wb


def _reorder(wb):
    order = ["Контент-план", "Размещения", "Бэклог", "Обновления", "Кейсы",
             "Отклонённые темы", "Дубли", "Нагрузка", "Источники сигналов",
             "Decision log", "Справочники"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)


# ------------------------------------------------------------ форматирование
def _apply_plan_formatting(ws, n_rows):
    if ws.max_row < 2:
        n_rows = 0
    last = max(2, ws.max_row)
    # score
    sc = _col_index(PLAN_COLUMNS, "score")
    if sc:
        L = get_column_letter(sc)
        rng = f"{L}2:{L}{last}"
        ws.conditional_formatting.add(rng, CellIsRule(operator="greaterThanOrEqual",
            formula=["72"], fill=FILL_GREEN))
        ws.conditional_formatting.add(rng, CellIsRule(operator="between",
            formula=["58", "71.999"], fill=FILL_YELLOW))
        ws.conditional_formatting.add(rng, CellIsRule(operator="lessThan",
            formula=["44"], fill=FILL_RED))
    # обязательные -> подсветка content_id
    mand = _col_index(PLAN_COLUMNS, "mandatory")
    cid = _col_index(PLAN_COLUMNS, "content_id")
    if mand and cid:
        ML = get_column_letter(mand)
        ws.conditional_formatting.add(
            f"A2:{get_column_letter(len(PLAN_COLUMNS))}{last}",
            FormulaRule(formula=[f'$%s2="да"' % ML], fill=FILL_MANDATORY, stopIfTrue=False))
    # риски
    risk = _col_index(PLAN_COLUMNS, "risk_flags")
    if risk:
        RL = get_column_letter(risk)
        ws.conditional_formatting.add(f"{RL}2:{RL}{last}",
            FormulaRule(formula=[f'LEN(TRIM({RL}2))>0'], fill=FILL_RISK, stopIfTrue=False))
    _add_plan_validation(ws, last)


def _add_plan_validation(ws, last):
    dvs = {
        "strategy_role": _ref_list("роли"),
        "funnel_stage": _ref_list("воронка"),
        "publication_kind": _ref_list("вид"),
        "effort": _ref_list("эфорт"),
        "decision_status": _ref_list("решение"),
        "mandatory": '"да,нет"',
        "lead_potential": '"высокий,средний,низкий"',
        "seo_potential": '"высокий,средний,низкий"',
        "product_priority": '"высокий,средний,низкий"',
    }
    for key, formula in dvs.items():
        idx = _col_index(PLAN_COLUMNS, key)
        if not idx:
            continue
        dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        ws.add_data_validation(dv)
        L = get_column_letter(idx)
        dv.add(f"{L}2:{L}{max(last, 2)}")


def _apply_placement_validation(ws, n_rows):
    last = max(2, ws.max_row)
    idx = _col_index(PLACEMENT_COLUMNS, "status")
    if idx:
        dv = DataValidation(type="list", formula1=_ref_list("статус_разм"), allow_blank=True)
        ws.add_data_validation(dv)
        L = get_column_letter(idx)
        dv.add(f"{L}2:{L}{last}")
    idx2 = _col_index(PLACEMENT_COLUMNS, "platform")
    if idx2:
        dv2 = DataValidation(type="list", formula1=_ref_list("площадки"), allow_blank=True)
        ws.add_data_validation(dv2)
        L2 = get_column_letter(idx2)
        dv2.add(f"{L2}2:{L2}{last}")


def _apply_dup_formatting(ws, n_rows):
    last = max(2, ws.max_row)
    rel = _col_index(DUP_COLUMNS, "relationship")
    if rel:
        L = get_column_letter(rel)
        ws.conditional_formatting.add(f"A2:{get_column_letter(len(DUP_COLUMNS))}{last}",
            FormulaRule(formula=[f'$%s2="дубль"' % L], fill=FILL_RED, stopIfTrue=False))
        ws.conditional_formatting.add(f"A2:{get_column_letter(len(DUP_COLUMNS))}{last}",
            FormulaRule(formula=[f'$%s2="каннибализация"' % L], fill=FILL_YELLOW, stopIfTrue=False))


# ------------------------------------------------------------ Справочники
# Диапазоны справочника для выпадающих списков.
_REF_LAYOUT = [
    ("роли", "Роли (Protect/Convert/Build/Prove)", schema.STRATEGY_ROLES),
    ("воронка", "Этапы воронки", schema.FUNNEL_STAGES),
    ("вид", "Вид публикации", schema.PUBLICATION_KINDS),
    ("эфорт", "Трудоёмкость", schema.EFFORT_LEVELS),
    ("решение", "Статус решения", schema.DECISION_STATUSES),
    ("статус_разм", "Статус размещения", schema.PLACEMENT_STATUSES),
    ("площадки", "Площадки", None),  # заполняется из таксономии в build
]
_REF_COL_INDEX: Dict[str, int] = {}


def _build_reference_sheet(ws, platforms: List[str] = None):
    ws.append(["Справочники контент-плана — используются для выпадающих списков"])
    ws["A1"].font = Font(bold=True, size=11)
    layout = []
    for key, title, values in _REF_LAYOUT:
        if key == "площадки":
            values = platforms or ["Блог Kaiten", "Habr", "VC.ru", "Бизнес-секреты",
                                   "CRMindex", "Startpack", "Sostav", "Partnerkin",
                                   "Tproger", "E-xecutive", "Внешняя (другое)"]
        layout.append((key, title, values))
    for col_i, (key, title, values) in enumerate(layout, start=1):
        _REF_COL_INDEX[key] = col_i
        L = get_column_letter(col_i)
        ws.column_dimensions[L].width = 34
        hc = ws.cell(row=2, column=col_i, value=title)
        hc.fill = HEADER_FILL
        hc.font = HEADER_FONT
        hc.alignment = HEADER_ALIGN
        for r, val in enumerate(values, start=3):
            ws.cell(row=r, column=col_i, value=val).font = CELL_FONT
    ws.freeze_panes = "A3"


def _ref_list(key: str) -> str:
    """Формула-ссылка на диапазон справочника (работает после открытия файла)."""
    col = _REF_COL_INDEX.get(key, 1)
    L = get_column_letter(col)
    return f"=Справочники!${L}$3:${L}$40"


# ------------------------------------------------------------ Нагрузка
def _build_load_sheet(ws, capacity_lines: List[Dict], cycle_info: Dict):
    ws.append(["Нагрузка редакции по неделям"])
    ws["A1"].font = Font(bold=True, size=12)
    if cycle_info:
        ws.append([f"Цикл: {cycle_info.get('slug','')} · целевая нагрузка "
                   f"{cycle_info.get('per_week','?')} материалов/нед · "
                   f"потолок {cycle_info.get('max_per_week','?')}/нед"])
    header_row = ws.max_row + 1
    cols = [("week", "Неделя", 30), ("count", "Материалов", 14),
            ("high_effort", "Тяжёлых", 12), ("flag", "Пометка", 40)]
    for i, (k, lbl, w) in enumerate(cols, start=1):
        c = ws.cell(row=header_row, column=i, value=lbl)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = HEADER_ALIGN
        ws.column_dimensions[get_column_letter(i)].width = w
    for line in capacity_lines:
        ws.append([line.get("week", ""), line.get("count", 0),
                   line.get("high_effort", 0), line.get("flag", "")])
    ws.freeze_panes = f"A{header_row+1}"


def save_workbook(wb: Workbook, path: str) -> str:
    import os
    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    return path
